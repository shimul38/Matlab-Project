"""
format_results.py
-----------------
Applies professional styling to QRS_Results.xlsx after MATLAB writes the raw data.

Run from the same folder as QRS_Results.xlsx:
    python format_results.py

Or from MATLAB console:
    >> system('python format_results.py');
"""

from openpyxl import load_workbook
from openpyxl.styles import (
    Font, PatternFill, Alignment, Border, Side, GradientFill
)
from openpyxl.styles.numbers import FORMAT_PERCENTAGE_00
from openpyxl.utils import get_column_letter
import os

INFILE  = 'QRS_Results.xlsx'
OUTFILE = 'QRS_Results.xlsx'   # overwrite in place

# ── Colour palette ──────────────────────────────────────────────────────────
C_HEADER_BG  = '1F3864'   # deep navy
C_HEADER_FG  = 'FFFFFF'   # white
C_SUBHDR_BG  = '2E75B6'   # mid blue (HRV columns)
C_ACC_BG     = 'D6E4F0'   # light blue accent (every other row)
C_GREEN_BG   = 'E2EFDA'   # light green (Se / Pp columns)
C_GREEN_FG   = '375623'   # dark green text
C_RED_BG     = 'FCE4D6'   # light red (FP / FN)
C_RED_FG     = '833C00'   # dark red text
C_GOLD_BG    = 'FFF2CC'   # gold (summary section headers)
C_GOLD_FG    = '7F6000'   # dark gold
C_WHITE      = 'FFFFFF'
C_BORDER     = 'BDD7EE'   # light blue border


def make_fill(hex_color):
    return PatternFill('solid', fgColor=hex_color)


def thin_border(color=C_BORDER):
    s = Side(style='thin', color=color)
    return Border(left=s, right=s, top=s, bottom=s)


def header_font(size=11, color=C_HEADER_FG):
    return Font(name='Calibri', bold=True, size=size, color=color)


def body_font(bold=False, color='000000', size=10):
    return Font(name='Calibri', bold=bold, size=size, color=color)


def center():
    return Alignment(horizontal='center', vertical='center', wrap_text=True)


def left():
    return Alignment(horizontal='left', vertical='center', wrap_text=False)


# ── Helper: auto-fit column widths ──────────────────────────────────────────
def autofit(ws, min_w=8, max_w=30):
    for col in ws.columns:
        max_len = 0
        col_letter = get_column_letter(col[0].column)
        for cell in col:
            try:
                if cell.value:
                    max_len = max(max_len, len(str(cell.value)))
            except Exception:
                pass
        ws.column_dimensions[col_letter].width = min(max(max_len + 2, min_w), max_w)


# ── FORMAT SHEET 1: Per_Record ───────────────────────────────────────────────
def format_per_record(ws):
    # Column index map (1-based)
    # 1:Record 2:Beats 3:BPM 4:SDNN 5:RMSSD 6:pNN50
    # 7:TP 8:FP 9:FN 10:Se 11:Pp 12:Diagnosis 13:Validated
    HRV_COLS   = {4, 5, 6}       # SDNN, RMSSD, pNN50
    BAD_COLS   = {8, 9}           # FP, FN
    GOOD_COLS  = {10, 11}         # Se, Pp
    MAX_COL    = ws.max_column
    MAX_ROW    = ws.max_row

    # Row 1: header
    for col in range(1, MAX_COL + 1):
        cell = ws.cell(row=1, column=col)
        if col in HRV_COLS:
            cell.fill      = make_fill(C_SUBHDR_BG)
        else:
            cell.fill      = make_fill(C_HEADER_BG)
        cell.font      = header_font()
        cell.alignment = center()
        cell.border    = thin_border('FFFFFF')

    # Row height for header
    ws.row_dimensions[1].height = 36

    # Data rows
    for row in range(2, MAX_ROW + 1):
        row_fill = make_fill(C_ACC_BG) if row % 2 == 0 else make_fill(C_WHITE)

        for col in range(1, MAX_COL + 1):
            cell = ws.cell(row=row, column=col)
            cell.border = thin_border()

            if col in GOOD_COLS:
                cell.fill      = make_fill(C_GREEN_BG)
                cell.font      = body_font(bold=True, color=C_GREEN_FG)
                cell.alignment = center()
                # Number format for percentages
                if cell.value not in (None, ''):
                    cell.number_format = '0.00'
            elif col in BAD_COLS:
                cell.fill      = make_fill(C_RED_BG)
                cell.font      = body_font(color=C_RED_FG)
                cell.alignment = center()
            elif col in HRV_COLS:
                cell.fill      = make_fill('EBF3FB')
                cell.font      = body_font()
                cell.alignment = center()
                if cell.value not in (None, ''):
                    cell.number_format = '0.0'
            elif col == 12:   # Diagnosis
                cell.fill      = row_fill
                cell.font      = body_font(bold=(str(cell.value) != 'Normal Sinus Rhythm'))
                cell.alignment = left()
            elif col == 13:   # Validated
                cell.fill      = row_fill
                cell.alignment = center()
                cell.font      = body_font()
                # Show TRUE/FALSE more clearly
                if cell.value is True:
                    cell.value = 'Yes'
                    cell.font  = body_font(bold=True, color=C_GREEN_FG)
                elif cell.value is False:
                    cell.value = 'No'
                    cell.font  = body_font(color=C_RED_FG)
            else:
                cell.fill      = row_fill
                cell.font      = body_font()
                cell.alignment = center()
                # Round numeric columns
                if col == 3 and cell.value not in (None, ''):   # BPM
                    cell.number_format = '0.0'

        ws.row_dimensions[row].height = 18

    # Freeze header row + Record column
    ws.freeze_panes = 'B2'

    # Auto-fit columns
    autofit(ws)

    # Add alternating column shading hint via named style is not needed;
    # autofit + row banding is enough visually.


# ── FORMAT SHEET 2: Summary ──────────────────────────────────────────────────
def format_summary(ws):
    MAX_ROW = ws.max_row

    for row in range(1, MAX_ROW + 1):
        cell_a = ws.cell(row=row, column=1)
        cell_b = ws.cell(row=row, column=2)

        val_a = str(cell_a.value) if cell_a.value else ''

        # Section header rows (start with ---)
        if val_a.startswith('---'):
            for c in (cell_a, cell_b):
                c.fill      = make_fill(C_GOLD_BG)
                c.font      = Font(name='Calibri', bold=True, size=11, color=C_GOLD_FG)
                c.alignment = left()
                c.border    = thin_border(C_GOLD_FG)
            ws.row_dimensions[row].height = 22
            continue

        # Main header row (row 1: Metric / Value)
        if row == 1:
            for c in (cell_a, cell_b):
                c.fill      = make_fill(C_HEADER_BG)
                c.font      = header_font()
                c.alignment = center()
                c.border    = thin_border('FFFFFF')
            ws.row_dimensions[row].height = 30
            continue

        # Empty separator rows
        if not cell_a.value and not cell_b.value:
            ws.row_dimensions[row].height = 8
            continue

        # Normal data rows — alternate shading
        row_fill = make_fill(C_ACC_BG) if row % 2 == 0 else make_fill(C_WHITE)
        cell_a.fill      = row_fill
        cell_b.fill      = row_fill
        cell_a.font      = body_font(bold=True)
        cell_a.alignment = left()
        cell_b.font      = body_font()
        cell_b.alignment = center()

        for c in (cell_a, cell_b):
            c.border = thin_border()

        # Highlight accuracy / Se / Pp rows
        if any(kw in val_a for kw in ('Sensitivity', 'Predictivity', 'Accuracy')):
            cell_b.fill = make_fill(C_GREEN_BG)
            cell_b.font = body_font(bold=True, color=C_GREEN_FG)
            if cell_b.value not in (None, ''):
                cell_b.number_format = '0.00'

        ws.row_dimensions[row].height = 18

    ws.freeze_panes = 'A2'
    ws.column_dimensions['A'].width = 32
    ws.column_dimensions['B'].width = 18


# ── MAIN ─────────────────────────────────────────────────────────────────────
def main():
    if not os.path.exists(INFILE):
        print(f'ERROR: {INFILE} not found. Run the MATLAB batch script first.')
        return

    print(f'Loading {INFILE}...')
    wb = load_workbook(INFILE)

    if 'Per_Record' in wb.sheetnames:
        print('  Formatting Per_Record sheet...')
        format_per_record(wb['Per_Record'])

    if 'Summary' in wb.sheetnames:
        print('  Formatting Summary sheet...')
        format_summary(wb['Summary'])

    # Set Per_Record as the active sheet on open
    wb.active = wb['Per_Record']

    wb.save(OUTFILE)
    print(f'Done! Styled workbook saved to: {OUTFILE}')


if __name__ == '__main__':
    main()
